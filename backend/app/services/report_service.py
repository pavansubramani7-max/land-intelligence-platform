import io
import uuid
from datetime import datetime
from typing import Dict, Any

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf_report(data: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=20,
                                  textColor=colors.HexColor("#1a365d"), alignment=TA_CENTER)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=14,
                                    textColor=colors.HexColor("#2d3748"))
    body_style = styles["Normal"]

    story.append(Paragraph("Land Intelligence Platform", title_style))
    story.append(Paragraph("Property Analysis Report", title_style))
    story.append(Spacer(1, 0.2 * inch))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1a365d")))
    story.append(Spacer(1, 0.2 * inch))

    meta = [
        ["Report ID:", data.get("report_id", str(uuid.uuid4())[:8])],
        ["Property:", data.get("property_address", "N/A")],
        ["Date:", datetime.now().strftime("%B %d, %Y")],
        ["Land ID:", str(data.get("land_id", "N/A"))],
    ]
    meta_table = Table(meta, colWidths=[2 * inch, 4 * inch])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("1. Valuation Summary", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    val_data = data.get("valuation", {})
    val_table_data = [
        ["Metric", "Value"],
        ["Estimated Value", f"₹{val_data.get('estimated_value', 0):,.2f}"],
        ["Confidence Score", f"{val_data.get('confidence_score', 0) * 100:.1f}%"],
        ["RF Prediction", f"₹{val_data.get('rf_prediction', 0):,.2f}"],
        ["XGB Prediction", f"₹{val_data.get('xgb_prediction', 0):,.2f}"],
        ["ANN Prediction", f"₹{val_data.get('ann_prediction', 0):,.2f}"],
    ]
    val_table = Table(val_table_data, colWidths=[3 * inch, 3 * inch])
    val_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(val_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("2. Dispute Risk Assessment", heading_style))
    dispute = data.get("dispute", {})
    dispute_data = [
        ["Risk Level", dispute.get("dispute_risk_label", "N/A").upper()],
        ["Risk Score", f"{dispute.get('dispute_risk_score', 0) * 100:.1f}%"],
        ["Risk Factors", ", ".join(dispute.get("risk_factors", [])) or "None"],
    ]
    dispute_table = Table(dispute_data, colWidths=[3 * inch, 3 * inch])
    dispute_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(dispute_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("3. Fraud Detection Results", heading_style))
    fraud = data.get("fraud", {})
    fraud_data = [
        ["Anomaly Detected", "YES" if fraud.get("is_anomaly") else "NO"],
        ["Fraud Probability", f"{fraud.get('fraud_probability', 0) * 100:.1f}%"],
        ["Anomaly Score", f"{fraud.get('anomaly_score', 0):.4f}"],
        ["Flags", ", ".join(fraud.get("fraud_flags", [])) or "None"],
    ]
    fraud_table = Table(fraud_data, colWidths=[3 * inch, 3 * inch])
    fraud_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(fraud_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("4. Price Forecast", heading_style))
    forecast = data.get("forecast", {})
    forecast_table_data = [
        ["Period", "Projected Value", "Growth Rate"],
        ["Current", f"₹{forecast.get('current_value', 0):,.2f}", "—"],
        ["1 Year", f"₹{forecast.get('forecast_1yr', 0):,.2f}", f"{forecast.get('growth_rate_1yr', 0):.1f}%"],
        ["3 Years", f"₹{forecast.get('forecast_3yr', 0):,.2f}", f"{forecast.get('growth_rate_3yr', 0):.1f}%"],
    ]
    forecast_table = Table(forecast_table_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    forecast_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d3748")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(forecast_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("5. Investment Recommendation", heading_style))
    rec = data.get("recommendation", {})
    rec_color = {"BUY": colors.green, "HOLD": colors.orange, "AVOID": colors.red}.get(
        rec.get("recommendation", "HOLD"), colors.grey
    )
    rec_style = ParagraphStyle("Rec", parent=styles["Normal"], fontSize=16,
                                textColor=rec_color, alignment=TA_CENTER)
    story.append(Paragraph(f"⬤ {rec.get('recommendation', 'N/A')}", rec_style))
    story.append(Paragraph(rec.get("reasoning", ""), body_style))
    story.append(Spacer(1, 0.5 * inch))

    disclaimer_style = ParagraphStyle("Disclaimer", parent=styles["Normal"], fontSize=8,
                                       textColor=colors.grey)
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        "DISCLAIMER: This report is generated by AI models and is for informational purposes only. "
        "It does not constitute legal or financial advice. Always consult a certified professional.",
        disclaimer_style,
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_excel_report(data: Dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")

    ws_val = wb.active
    ws_val.title = "Valuation"
    val_data = data.get("valuation", {})
    headers = ["Metric", "Value"]
    rows = [
        ["Estimated Value", val_data.get("estimated_value", 0)],
        ["Confidence Score", val_data.get("confidence_score", 0)],
        ["RF Prediction", val_data.get("rf_prediction", 0)],
        ["XGB Prediction", val_data.get("xgb_prediction", 0)],
        ["ANN Prediction", val_data.get("ann_prediction", 0)],
    ]
    for col, h in enumerate(headers, 1):
        cell = ws_val.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws_val.cell(row=row_idx, column=col_idx, value=val)

    ws_dispute = wb.create_sheet("Dispute Risk")
    dispute = data.get("dispute", {})
    for col, h in enumerate(["Metric", "Value"], 1):
        cell = ws_dispute.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, (k, v) in enumerate([
        ("Risk Label", dispute.get("dispute_risk_label", "")),
        ("Risk Score", dispute.get("dispute_risk_score", 0)),
    ], 2):
        ws_dispute.cell(row=row_idx, column=1, value=k)
        ws_dispute.cell(row=row_idx, column=2, value=v)

    ws_forecast = wb.create_sheet("Forecast")
    forecast = data.get("forecast", {})
    for col, h in enumerate(["Period", "Value", "Growth %"], 1):
        cell = ws_forecast.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, (period, val, growth) in enumerate([
        ("Current", forecast.get("current_value", 0), 0),
        ("1 Year", forecast.get("forecast_1yr", 0), forecast.get("growth_rate_1yr", 0)),
        ("3 Years", forecast.get("forecast_3yr", 0), forecast.get("growth_rate_3yr", 0)),
    ], 2):
        ws_forecast.cell(row=row_idx, column=1, value=period)
        ws_forecast.cell(row=row_idx, column=2, value=val)
        ws_forecast.cell(row=row_idx, column=3, value=growth)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
